import os
import json
from openpyxl import load_workbook


def parse_and_convert_to_json(directory):
    data_concretn = {}

    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)
            parse_xlsx_and_convert_to_json(file_path, directory, filename, data_concretn)

    # Создаем JSON-файл для data_concretn
    with open("data_concretn.json", 'w', encoding='utf-8') as json_file:
        json.dump(data_concretn, json_file, ensure_ascii=False, indent=4)


def parse_xlsx_and_convert_to_json(file_path, output_directory, filename, data_concretn):
    try:
        wb = load_workbook(file_path)
        json_data = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            json_data[sheet_name.rstrip()] = []

            for row in range(2, sheet.max_row + 1):
                data_entry = {}
                data_entry["header_date"] = str(sheet.cell(row=1, column=1).value)
                data_entry["header_pair"] = str(sheet.cell(row=1, column=2).value)

                room_a = str(sheet.cell(row=row, column=1).value).replace(".0", "").lower()
                room_d = str(sheet.cell(row=row, column=4).value).replace(".0", "").lower()
                room_a = str(sheet.cell(row=row, column=1).value).replace(".0", "").lower()
                room_d = str(sheet.cell(row=row, column=4).value).replace(".0", "").lower()
                group_b = str(sheet.cell(row=row, column=2).value).rstrip().lower()
                group_e = str(sheet.cell(row=row, column=5).value).rstrip().lower()
                teacher_c = str(sheet.cell(row=row, column=3).value).rstrip().lower()
                teacher_f = str(sheet.cell(row=row, column=6).value).rstrip().lower()

                if room_a:
                    if group_b == None or group_b == "none":
                        group_b = "Отсутствует"
                    if teacher_c == None or teacher_c == "none":
                        teacher_c = "Отсутствует"

                    entry = {"header_date": data_entry["header_date"], "header_pair": data_entry["header_pair"],
                             "room": room_a, "group": group_b, "teacher": teacher_c}
                    json_data[sheet_name.rstrip()].append(entry)
    # 040124 удалила деление на столбики из таблицы. а нахер я их делала вообще? ты дурак????
                if room_d:
                    if group_e == None or group_e == "none":
                        group_e = "Отсутствует"
                    if teacher_f == None or teacher_f == "none":
                        teacher_f = "Отсутствует"

                    entry = {"header_date": data_entry["header_date"], "header_pair": data_entry["header_pair"],
                             "room": room_d, "group": group_e, "teacher": teacher_f}
                    json_data[sheet_name.rstrip()].append(entry)

            date_key = os.path.splitext(filename)[0] + ".xlsx"
            if date_key not in data_concretn:
                data_concretn[date_key] = [sheet_name.rstrip()]
            else:
                data_concretn[date_key].append(sheet_name.rstrip())

            json_filename = f"{os.path.splitext(filename)[0]}.json"
            json_filepath = os.path.join(output_directory, json_filename)

            with open(json_filepath, 'w', encoding='utf-8') as json_file:
                json.dump(json_data, json_file, ensure_ascii=False, indent=4)

        # Создаем JSON-файл только с датами и названиями листов
        data_concretn_filename = "data_concretn.json"
        data_concretn_filepath = os.path.join(output_directory, data_concretn_filename)

        with open(data_concretn_filepath, 'w', encoding='utf-8') as json_file:
            json.dump(data_concretn, json_file, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"An error occurred: {e}")