import os
from openpyxl import load_workbook
from datetime import datetime


def parse_xlsx(file_path, search_date, search_pair, search_type, search_value):
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        if str(search_pair) in sheet_name:
            sheet = wb[sheet_name]

            header_date = sheet.cell(row=1, column=1).value
            header_pair = sheet.cell(row=1, column=2).value
            occupancy_cell = sheet.cell(row=1, column=3)

            occupancy_str = str(occupancy_cell.value)

            parts = occupancy_str.split(" корпус ") if " корпус " in occupancy_str else ["Нет данных", "Нет данных"]
            occupancy, percentage = parts[0], parts[1].split("% заполненность ")[0] if len(parts) > 1 else "Нет данных"

            if header_date != search_date:
                continue

            found = False

            for row in range(2, sheet.max_row + 1):
                room_a = str(sheet.cell(row=row, column=1).value)
                room_d = str(sheet.cell(row=row, column=4).value)
                group_b = str(sheet.cell(row=row, column=2).value)
                group_e = str(sheet.cell(row=row, column=5).value)
                teacher_c = str(sheet.cell(row=row, column=3).value)
                teacher_f = str(sheet.cell(row=row, column=6).value)

                if search_type == "кабинет":
                    if search_value == room_a:
                        print_details(occupancy, header_date, header_pair, room_a, group_b, teacher_c)
                        found = True

                    if search_value == room_d:
                        print_details(occupancy, header_date, header_pair, room_d, group_e, teacher_f)
                        found = True

                elif search_type == "группа":
                    if search_value == group_b:
                        print_details(occupancy, header_date, header_pair, room_a, group_b, teacher_c)
                        found = True

                    if search_value == group_e:
                        print_details(occupancy, header_date, header_pair, room_d, group_e, teacher_f)
                        found = True

                elif search_type == "преподаватель":
                    if search_value == teacher_c:
                        print_details(occupancy, header_date, header_pair, room_a, group_b, teacher_c)
                        found = True

                    if search_value == teacher_f:
                        print_details(occupancy, header_date, header_pair, room_d, group_e, teacher_f)
                        found = True

            if not found:
                print(f"Нет данных для типа {search_type} '{search_value}' на листе '{sheet_name}' в файле {file_path}")


def print_details(occupancy, header_date, header_pair, room, group, teacher):
    print()
    print("Корпус:", occupancy)
    print("Дата:", header_date)
    print("Пара:", header_pair)
    print("Кабинет:", room.replace('.0', '', 1) if room and any(c.isdigit() for c in room) else "отсутствует")
    print(f"Группа: {group}" if group is not None and group != "None" else "Группа: отсутствует")
    print(f"Преподаватель: {teacher}" if teacher is not None and teacher != "None" else "Преподаватель: отсутствует")


directory = "all_planchette"

search_date_str = input("Введи дату в виде дд.мм.гггг: ")
search_date = datetime.strptime(search_date_str, "%d.%m.%Y")
search_pair = input("Введи номер пары: ")
search_type = input("Выбери тип ввода (кабинет/группа/преподаватель): ")
search_value = str(input(f"Введи {search_type}: "))

if search_value.isdigit():
    search_value = f"{int(search_value)}.0"

for filename in os.listdir(directory):
    if filename.endswith(".xlsx") and search_date_str in filename:
        file_path = os.path.join(directory, filename)
        print("Обрабатываем файл:", file_path)
        parse_xlsx(file_path, search_date, search_pair, search_type, search_value)
